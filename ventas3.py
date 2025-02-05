import importlib.util
import os
import pandas as pd
import re
from tkinter import filedialog, Tk, Label, messagebox  # Importar messagebox para mostrar mensajes de error
from tkinter import ttk  # Importar ttk para una interfaz más moderna
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def obtener_tipo_y_numero(texto):
    """
    Extrae el tipo de documento y número desde un texto.
    """
    if pd.isna(texto):
        return None, None
    # Buscar el tipo de documento y número
    match = re.search(r'(.*?)\s*/\s*(\d+)', texto)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return None, None

def limpiar_filas(df, columna='SKU'):
    """
    Limpia y formatea una columna del DataFrame de manera segura.
    """
    df_copy = df.copy()
    # Convert column to string type first
    df_copy[columna] = df_copy[columna].astype(str)
    # Apply cleaning operations
    df_copy[columna] = (df_copy[columna]
                    .str.replace(r'\.0$', '', regex=True)
                    .str.replace(r'^="|"$', '', regex=True)
                    .str.strip()
                    .str.lower())
    return df_copy

def Costos_negativos(df1, columna='Tipo Movimiento', fila='Boleta', columna_a_cambiar='Costo'):
    df1.loc[df1[columna] == fila, columna_a_cambiar] = -df1.loc[df1[columna] == fila, columna_a_cambiar]
    return df1

def calcular_costos_totales(df, sku_col, costo_neto_col, costo_total_col):
    """
    Calcula el costo total por SKU y lo añade como una nueva columna en el DataFrame.
    """
    df[costo_total_col] = df.groupby(sku_col)[costo_neto_col].transform("sum")
    return df

def crear_diccionarios_mapeo(df, sku_col, costo_neto_col, costo_total_col):
    """
    Crea diccionarios para mapear SKU con costos unitarios y totales.
    """
    costo_unitario = df.drop_duplicates(sku_col).set_index(sku_col)[costo_neto_col].to_dict()
    costo_total = df.drop_duplicates(sku_col).set_index(sku_col)[costo_total_col].to_dict()
    return costo_unitario, costo_total

def aplicar_condicion(df, condicion_col, condicion_valor):
    """
    Crea una condición booleana para filtrar las filas de un DataFrame.
    """
    return df[condicion_col].str.lower() == condicion_valor.lower()

def asignar_costos_a_df(df, condicion, sku_col, costo_unitario, costo_total, costo_neto_col, costo_total_col):
    """
    Asigna costos netos unitarios y totales a un DataFrame basado en una condición y diccionarios de mapeo.
    """
    df.loc[condicion, costo_neto_col] = df.loc[condicion, sku_col].map(costo_unitario)
    df.loc[condicion, costo_total_col] = df.loc[condicion, sku_col].map(costo_total)
    return df

def asignar_costos(df1, df2, tipo_documento_col="Tipo de Documento", sku_col="SKU", 
                costo_neto_col="Costo Neto", costo_total_col="Costo Total Neto", 
                condicion_valor="nota de crédito electrónica"):
    """
    Orquesta las funciones para calcular y asignar costos netos unitarios y totales a un DataFrame.
    """
    # Calcular costos totales en df2
    df2 = calcular_costos_totales(df2, sku_col, costo_neto_col, costo_total_col)

    # Crear los diccionarios de mapeo
    costo_unitario, costo_total = crear_diccionarios_mapeo(df2, sku_col, costo_neto_col, costo_total_col)

    # Crear la condición en df1
    condicion = aplicar_condicion(df1, tipo_documento_col, condicion_valor)

    # Asignar los costos al DataFrame
    df1 = asignar_costos_a_df(df1, condicion, sku_col, costo_unitario, costo_total, costo_neto_col, costo_total_col)

    return df1

def agregar_cantidad_df2_a_df1(df1, df2, sku_col="SKU"):
    """
    Agrega una columna 'Cantidad' en df1 con la cantidad de veces que se repite cada SKU en df2.

    Parámetros:
    - df1: DataFrame al que se le agregará la columna 'Cantidad'.
    - df2: DataFrame desde el cual se contarán los SKUs.
    - sku_col: Nombre de la columna que contiene los SKUs (por defecto 'SKU').
    
    Devuelve:
    - df1 con una nueva columna 'Cantidad'.
    """
    # Contar la cantidad de repeticiones de cada SKU en df2
    sku_to_cantidad = df2[sku_col].value_counts().to_dict()
    # Mapear los valores al DataFrame 1
    df1['Cantidad'] = df1[sku_col].map(sku_to_cantidad)

    # Rellenar valores NaN con 0
    df1['Cantidad'] = df1['Cantidad'].fillna(0).astype(int)

    return df1

def calcular_margen(ventas, costos):
    """Calcula el margen de forma segura evitando división por cero."""
    try:
        return (ventas - abs(costos)) / ventas * 100
    except ZeroDivisionError:
        return 0

def Contador_Saltador_de_lineas_hasta(df, columna, valor):
    """
    Cuenta las filas que se deben saltar hasta encontrar un valor específico en una columna.
    """
    return df[columna].eq(valor).idxmax() + 1


def filtrar_y_crear_tabla_dinamica(df, condiciones, index_column, aggfunc_dict):
    """
    Filtra un DataFrame basado en condiciones y luego crea una tabla dinámica.
    
    Parameters:
    - df: DataFrame de entrada a filtrar.
    - condiciones: diccionario donde las claves son los nombres de las columnas
        y los valores son las condiciones de filtrado.
    - index_column: Columna sobre la que se crea la tabla dinámica (índice).
    - aggfunc_dict: Diccionario que define las funciones de agregación (por ejemplo, suma, promedio).
    
    Returns:
    - DataFrame filtrado y con tabla dinámica creada.
    """
    # Aplicar las condiciones de filtrado
    for columna, condicion in condiciones.items():
        if isinstance(condicion, list):
            df = df[df[columna].isin(condicion)]  # Filtrar por lista de valores
        elif callable(condicion):
            df = df[df[columna].apply(condicion)]  # Filtrar por función
        else:
            df = df[df[columna] == condicion]  # Filtrar por valor específico
    
    # Crear la tabla dinámica
    df_pivot = df.pivot_table(index=index_column, aggfunc=aggfunc_dict)
    
    return df, df_pivot


def crear_condiciones_no_igual(columnas, valores):
    """
    Crea un diccionario de condiciones dinámicas para filtrar un DataFrame basado únicamente en `!=`.

    Parameters:
    - columnas: lista de nombres de columnas en el DataFrame a filtrar.
    - valores: lista de valores o listas de valores que deben excluirse.

    Returns:
    - Diccionario de condiciones que puede ser usado con una función de filtrado dinámico.
    """
    if len(columnas) != len(valores):
        raise ValueError("Las listas 'columnas' y 'valores' deben tener la misma longitud.")

    # Crear condiciones usando `!=`
    condiciones = {}
    for columna, valor in zip(columnas, valores):
        condiciones[columna] = lambda x, v=valor: x not in v if isinstance(v, list) else x != v

    return condiciones

# Variables globales - eliminar columna_no_considerada_entry
ruta_ventas = None
ruta_costos = None
ruta_ano = None
lineas_ventas = None
lineas_ano = None
label_ventas = None
label_costos = None
label_ano = None
fila_ventas_entry = None
fila_ano_entry = None

# Agregar variables globales
filtros_columnas = []
filtros_valores = []

def actualizar_etiqueta(label, texto):
    """Actualiza el texto de una etiqueta de manera segura"""
    if label:
        label.config(text=texto)

def seleccionar_archivo_ventas():
    global ruta_ventas
    ruta_ventas = filedialog.askopenfilename(title="Selecciona el archivo de ventas", filetypes=[("Archivos Excel", "*.xlsx")])
    if ruta_ventas:
        actualizar_etiqueta(label_ventas, "Archivo seleccionado: " + os.path.basename(ruta_ventas))

def seleccionar_archivo_costos():
    global ruta_costos
    ruta_costos = filedialog.askopenfilename(title="Selecciona el archivo de costos", filetypes=[("Archivos Excel", "*.xls")])
    if ruta_costos:
        actualizar_etiqueta(label_costos, "Archivo seleccionado: " + os.path.basename(ruta_costos))

def seleccionar_archivo_ano():
    global ruta_ano
    ruta_ano = filedialog.askopenfilename(title="Selecciona el archivo del año", filetypes=[("Archivos Excel", "*.xlsx")])
    if ruta_ano:
        actualizar_etiqueta(label_ano, "Archivo seleccionado: " + os.path.basename(ruta_ano))

def seleccionar_lineas():
    global lineas_ventas, lineas_ano
    try:
        if not ruta_ventas:
            messagebox.showerror("Error", "Debe seleccionar el archivo de ventas.")
            return

        # Validar fila de inicio para ventas mensuales
        lineas_ventas = fila_ventas_entry.get().strip()
        if not lineas_ventas.isdigit():
            messagebox.showerror("Error", "Por favor, ingrese un número válido para la fila de inicio de ventas mensuales.")
            return
        lineas_ventas = int(lineas_ventas)

        # Validar fila de inicio para ventas anuales
        if ruta_ano:
            lineas_ano = fila_ano_entry.get().strip()
            if not lineas_ano.isdigit():
                messagebox.showerror("Error", "Por favor, ingrese un número válido para la fila de inicio de ventas anuales.")
                return
            lineas_ano = int(lineas_ano)

        messagebox.showinfo("Éxito", "Líneas seleccionadas correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error: {e}")

def seleccionar_filtros_tablas():
    """Función para seleccionar los filtros de las tablas dinámicas"""
    global filtros_columnas, filtros_valores
    try:
        # Obtener columnas y valores de los entries
        columnas = filtros_columnas_entry.get().strip()
        valores = filtros_valores_entry.get().strip()
        
        if columnas and valores:
            filtros_columnas = [col.strip() for col in columnas.split(',')]
            filtros_valores = [val.strip().split(';') for val in valores.split(',')]
            
            if len(filtros_columnas) != len(filtros_valores):
                raise ValueError("El número de columnas debe coincidir con el número de grupos de valores")
            
            messagebox.showinfo("Éxito", "Filtros para tablas dinámicas seleccionados correctamente.")
        else:
            filtros_columnas = []
            filtros_valores = []
            messagebox.showinfo("Info", "No se aplicarán filtros.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al seleccionar filtros: {e}")

def ejecutar_programa():
    try:
        if not ruta_ventas:
            messagebox.showerror("Error", "Debe seleccionar el archivo de ventas.")
            return
        if lineas_ventas is None:
            messagebox.showerror("Error", "Debe seleccionar las líneas de inicio antes de ejecutar.")
            return

        # Leer los archivos
        df1 = pd.read_excel(ruta_ventas, skiprows=lineas_ventas - 1, header=0)
        df2 = pd.read_html(ruta_costos, header=0)[0] if ruta_costos else None
        df3 = pd.read_excel(ruta_ano, skiprows=lineas_ano - 1, header=0) if ruta_ano else None

        # Proceso de análisis o manipulación
        # Aquí puedes añadir la lógica para manipular y guardar los DataFrames


        # Eliminar ultima linea y quitar.0
        df1 = df1.iloc[:-1]

        # Añadir cantidades en nota de crédito electrónica y que estas sean negativas
        df1 = limpiar_filas(df1, 'SKU')
        df1['Producto']=df1['SKU'].astype(str).str[:8]
        if df2 is not None:
            df2 = limpiar_filas(df2, 'SKU')
            diccionario = dict(zip(df2['SKU'], df2['Costo Neto']))
            conteo_skus = df2['SKU'].value_counts()
            condicion = df1['Tipo de Documento'] == 'nota de crédito electrónica'
            df1.loc[condicion, "Cantidad"] = df1.loc[condicion, "SKU"].map(conteo_skus)
            df1.loc[df1['Tipo de Documento'] == 'nota de crédito electrónica', 'Cantidad'] = -abs(df1.loc[df1['Tipo de Documento'] == 'nota de crédito electrónica', 'Cantidad'])

            # Añadir costos unitarios y costos totales a df1 
            df1['Costo Neto Unitario'] = df1['SKU'].map(diccionario).fillna(0)
            df1.loc[df1['Tipo de Documento'] == 'nota de crédito electrónica', 'Costo Total Neto'] = df1['Costo Neto Unitario'] * df1['Cantidad']

            # Costos para las boletas son negativos
            df1 = Costos_negativos(df1, 'Tipo de Documento', 'boleta electrónica t', 'Costo Neto Unitario')
            df1 = Costos_negativos(df1, 'Tipo de Documento', 'boleta electrónica t', 'Costo Total Neto')
            df1 = Costos_negativos(df1, 'Tipo de Documento', 'nota de crédito electrónica', 'Costo Total Neto')


        # Quitar .0 de Costo Total Neto y Costo Neto Unitario
        df1['Costo Total Neto'] = df1['Costo Total Neto'].fillna(0).round(0).astype('int64')
        df1['Costo Neto Unitario'] = df1['Costo Neto Unitario'].fillna(0).round(0).astype('int64')

    
        # Crear columna Total venta antes de crear las tablas dinámicas
        df1['Total venta'] = df1['Subtotal Bruto'] - df1['Subtotal Impuestos']

        # Ahora con excel año 
        if df3 is not None:

            df3 = limpiar_filas(df3, 'Nº Documento')
            df1 = limpiar_filas(df1, 'Numero Documento')

            # Solo considerar notas de crédito electrónica en df3
            # Solo Boletas 
            Boletas = df3.loc[aplicar_condicion(df3, 'Tipo Documento', 'BOLETA ELECTRÓNICA T')]

            # 1. Procesar Boletas
            condicion_mayorista = (Boletas['Sucursal'] == 'SUCURSAL MAYORISTA')
            Boletas.loc[condicion_mayorista, 'Marketplace'] = 'T.NAVIDEÑA'
            
            # Mapear Boletas a df1
            df1["Numero Documento"] = df1["Numero Documento"].astype(str).str.strip()
            Boletas["Nº Documento"] = Boletas["Nº Documento"].astype(str).str.strip()
            NDocumento_to_marketplace = Boletas.set_index("Nº Documento")["Marketplace"].to_dict()
            condicion_boleta = df1['Tipo de Documento'].str.lower() == 'boleta electrónica t'
            df1.loc[condicion_boleta, "Marketplace"] = df1.loc[condicion_boleta, "Numero Documento"].map(NDocumento_to_marketplace)

            # 2. Procesar Facturas de Exportación
            fila_exportacion = Contador_Saltador_de_lineas_hasta(df3, 'Tipo Documento', 'FACTURA DE EXPORTACIÓN ELECTRÓNICA')
            Exportacion = pd.read_excel(ruta_ano, skiprows=fila_exportacion + lineas_ano - 2, header=0, usecols=lambda column: not column.startswith('Unnamed:'))
            Exportacion = Exportacion[Exportacion['Tipo Documento'] == 'FACTURA DE EXPORTACIÓN ELECTRÓNICA']            
            condicion_exportacion = (Exportacion['Sucursal'] == 'SUCURSAL MAYORISTA')

            # Aplicar cvd en Marketplace para facturas de exportacion 
            Exportacion['Marketplace'] = 'cvd'
            
            # Mapear Exportación a df1
            Exportacion["Nº Documento"] = Exportacion["Nº Documento"].astype(str).str.strip()
            NDocumento_to_exportacion = Exportacion.set_index("Nº Documento")["Marketplace"].to_dict()
            condicion_exportacion = df1['Tipo de Documento'].str.contains('FACTURA DE EXPORTACIÓN ELECTRÓNICA', case=False)
            df1.loc[condicion_exportacion, "Marketplace"] = df1.loc[condicion_exportacion, 'Numero Documento'].map(NDocumento_to_exportacion)

            # 3. Procesar Guías de Despacho
            fila_DESPACHO = Contador_Saltador_de_lineas_hasta(df3, 'Tipo Documento', 'GUÍA DE DESPACHO ELECTRÓNICA')
            DESPACHO = pd.read_excel(ruta_ano, skiprows=fila_DESPACHO + lineas_ano - 2, header=0, usecols=lambda column: not column.startswith('Unnamed:'))
            DESPACHO = DESPACHO[DESPACHO['Tipo Documento'] == 'GUÍA DE DESPACHO ELECTRÓNICA']
            DESPACHO["Nº Documento"] = DESPACHO["Nº Documento"].astype(str).str.strip()

            # 5. Procesar Facturas Electrónicas
            fila_facturas = Contador_Saltador_de_lineas_hasta(df3, 'Tipo Documento', 'FACTURA ELECTRÓNICA')
            facturas = pd.read_excel(ruta_ano, skiprows=fila_facturas + lineas_ano - 2, header=0, usecols=lambda column: not column.startswith('Unnamed:'))
            facturas = facturas[facturas['Tipo Documento'] == 'FACTURA ELECTRÓNICA']
            
            # Inicializar el resto como 'Sin datos'
            facturas['Marketplace'] = 'Sin datos'

            # Luego continuar con el proceso de documentos relacionados
            facturas[['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']] = pd.DataFrame(
                facturas['Documentos Relacionados (Tipo Documento / Nº Documento / Fecha Emisión / Monto del documento )']
                .apply(obtener_tipo_y_numero)
                .tolist(), columns=['Tipo', 'Numero']
            )

            # --- Optimización para actualizar 'Marketplace' en facturas ---
            # Eliminar el bucle for y usar mapeos vectorizados:
            map_despacho = DESPACHO.set_index('Nº Documento')['Marketplace'].to_dict()
            map_boletas   = Boletas.set_index('Nº Documento')['Marketplace'].to_dict()
            map_export    = Exportacion.set_index('Nº Documento')['Marketplace'].to_dict()

            facturas[['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']] = pd.DataFrame(
                facturas['Documentos Relacionados (Tipo Documento / Nº Documento / Fecha Emisión / Monto del documento )']
                .apply(obtener_tipo_y_numero).tolist(), columns=['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']
            )
            mask_despacho = facturas['Tipo_Doc_Relacionado'].str.contains('GUÍA DE DESPACHO ELECTRÓNICA', na=False)
            mask_boleta   = facturas['Tipo_Doc_Relacionado'].str.contains('BOLETA ELECTRÓNICA', na=False)
            mask_export   = facturas['Tipo_Doc_Relacionado'].str.contains('FACTURA DE EXPORTACIÓN ELECTRÓNICA', na=False)

            facturas.loc[mask_despacho, 'Marketplace'] = facturas.loc[mask_despacho, 'Num_Doc_Relacionado'].map(map_despacho)
            facturas.loc[mask_boleta,   'Marketplace'] = facturas.loc[mask_boleta, 'Num_Doc_Relacionado'].map(map_boletas)
            facturas.loc[mask_export,   'Marketplace'] = facturas.loc[mask_export, 'Num_Doc_Relacionado'].map(map_export)

            # Asignar marketplace basado en la sucursal después de procesar los documentos relacionados
            condicion_mayorista = (facturas['Sucursal'] == 'SUCURSAL MAYORISTA')
            condicion_venta_online2 = (facturas['Vendedor'] == 'VENTA ONLINE')
            facturas.loc[condicion_mayorista, 'Marketplace'] = 'vpm'
            facturas.loc[condicion_venta_online2, 'Marketplace'] = 'fcom'
            

            # Mapear Marketplace para facturas
            df1['Numero Documento'] = df1['Numero Documento'].astype(str).str.strip()
            facturas['Nº Documento'] = facturas['Nº Documento'].astype(str).str.strip()
            diccionario_facturas = dict(zip(facturas['Nº Documento'], facturas['Marketplace']))
            condicion_factura = df1['Tipo de Documento'].str.contains('FACTURA ELECTRÓNICA', case=False)
            df1.loc[condicion_factura, 'Marketplace'] = df1.loc[condicion_factura, 'Numero Documento'].map(diccionario_facturas)




            # 4. Procesar notas de crédito electrónica
            fila_Notas = Contador_Saltador_de_lineas_hasta(df3, 'Tipo Documento', 'NOTA DE CRÉDITO ELECTRÓNICA')
            notas = pd.read_excel(ruta_ano, skiprows=fila_Notas + lineas_ano - 2, header=0, usecols=lambda column: not column.startswith('Unnamed:'))
            notas = notas[notas['Tipo Documento'] == 'NOTA DE CRÉDITO ELECTRÓNICA']

            # Inicializar columna Marketplace solo una vez
            notas['Marketplace'] = 'Sin datos'

            # Aplicar condiciones para asignar Marketplace en orden específico
            # Primero VENTA ONLINE
            condicion_venta_online = (notas['Vendedor'] == 'VENTA ONLINE')


            # Luego SUCURSAL MAYORISTA (solo si no es VENTA ONLINE)
            condicion_mayorista = (notas['Sucursal'] == 'SUCURSAL MAYORISTA') & (~condicion_venta_online)


            # Extraer tipo de documento y número
            notas[['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']] = pd.DataFrame(
                notas['Documentos Relacionados (Tipo Documento / Nº Documento / Fecha Emisión / Monto del documento )']
                .apply(obtener_tipo_y_numero)
                .tolist(), columns=['Tipo', 'Numero']
            )

            # Se puede aplicar enfoque similar para 'notas' y 'debitos'
            # Por ejemplo, para 'notas':
            map_notas = {
                'Despacho': DESPACHO.set_index('Nº Documento')['Marketplace'].to_dict(),
                'Boleta':   Boletas.set_index('Nº Documento')['Marketplace'].to_dict(),
                'Export':   Exportacion.set_index('Nº Documento')['Marketplace'].to_dict(),
                'Factura':  facturas.set_index('Nº Documento')['Marketplace'].to_dict()
            }
            notas[['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']] = pd.DataFrame(
                notas['Documentos Relacionados (Tipo Documento / Nº Documento / Fecha Emisión / Monto del documento )']
                .apply(obtener_tipo_y_numero).tolist(), columns=['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']
            )
            # Actualizar Marketplace en notas según cada condición de forma vectorizada:
            mask_despacho_notas = notas['Tipo_Doc_Relacionado'].str.contains('GUÍA DE DESPACHO ELECTRÓNICA', na=False)
            mask_boleta_notas   = notas['Tipo_Doc_Relacionado'].str.contains('BOLETA ELECTRÓNICA', na=False)
            mask_export_notas   = notas['Tipo_Doc_Relacionado'].str.contains('FACTURA DE EXPORTACIÓN ELECTRÓNICA', na=False)
            mask_factura_notas  = notas['Tipo_Doc_Relacionado'].str.contains('FACTURA ELECTRÓNICA', na=False)

            notas.loc[mask_despacho_notas, 'Marketplace'] = notas.loc[mask_despacho_notas, 'Num_Doc_Relacionado'].map(map_notas['Despacho'])
            notas.loc[mask_boleta_notas,   'Marketplace'] = notas.loc[mask_boleta_notas, 'Num_Doc_Relacionado'].map(map_notas['Boleta'])
            notas.loc[mask_export_notas,   'Marketplace'] = notas.loc[mask_export_notas, 'Num_Doc_Relacionado'].map(map_notas['Export'])
            notas.loc[mask_factura_notas,  'Marketplace'] = notas.loc[mask_factura_notas, 'Num_Doc_Relacionado'].map(map_notas['Factura'])

            notas['Marketplace'] = notas['Marketplace'].fillna('Sin datos')
            condicion_mayorista1 = (notas['Sucursal'] == 'SUCURSAL MAYORISTA') & (notas['Marketplace'] == 'Sin datos')
            notas.loc[condicion_mayorista1, 'Marketplace'] = 'vpm'
            condicion_venta_online3 = (notas['Vendedor'] == 'VENTA ONLINE') & (notas['Sucursal'] == 'Casa Matriz') & (notas['Cliente'] == 'maestro')
            notas.loc[condicion_venta_online3, 'Marketplace'] = 'fcom'            
            # Mapear Marketplace para notas de credito
            df1['Numero Documento'] = df1['Numero Documento'].astype(str).str.strip()
            notas['Nº Documento'] = notas['Nº Documento'].astype(str).str.strip()
            diccionario_notas = dict(zip(notas['Nº Documento'], notas['Marketplace']))
            condicion_notas = df1['Tipo de Documento'].str.contains('NOTA DE CRÉDITO ELECTRÓNICA', case=False)
            df1.loc[condicion_notas, 'Marketplace'] = df1.loc[condicion_notas, 'Numero Documento'].map(diccionario_notas)



            #Para NOTA DE DÉBITO ELECTRÓNICA
            fila_debito = Contador_Saltador_de_lineas_hasta(df3, 'Tipo Documento', 'NOTA DE DÉBITO ELECTRÓNICA')
            debitos = pd.read_excel(ruta_ano, skiprows=fila_debito + lineas_ano - 2, header=0, usecols=lambda column: not column.startswith('Unnamed:'))
            debitos = debitos[debitos['Tipo Documento'] == 'NOTA DE DÉBITO ELECTRÓNICA']

            # Extraer tipo de documento y número para notas de débito
            debitos[['Tipo_Doc_Relacionado', 'Num_Doc_Relacionado']] = pd.DataFrame(
                debitos['Documentos Relacionados (Tipo Documento / Nº Documento / Fecha Emisión / Monto del documento )']
                .apply(obtener_tipo_y_numero)
                .tolist(), columns=['Tipo', 'Numero']
            )

            # Inicializar columna Marketplace
            debitos['Marketplace'] = 'Sin datos'

            # Procesar cada tipo de documento en notas de débito
            for idx, debito in debitos.iterrows():
                tipo_doc = debito['Tipo_Doc_Relacionado']
                num_doc = debito['Num_Doc_Relacionado']
                
                if pd.notna(tipo_doc) and pd.notna(num_doc):
                    if 'GUÍA DE DESPACHO ELECTRÓNICA' in tipo_doc:
                        if num_doc in DESPACHO.set_index('Nº Documento').index:
                            debitos.at[idx, 'Marketplace'] = DESPACHO.loc[DESPACHO['Nº Documento'] == num_doc, 'Marketplace'].iloc[0]
                    elif 'FACTURA ELECTRÓNICA' in tipo_doc:
                        if num_doc in facturas.set_index('Nº Documento').index:
                            debitos.at[idx, 'Marketplace'] = facturas.loc[facturas['Nº Documento'] == num_doc, 'Marketplace'].iloc[0]
                    elif 'BOLETA ELECTRÓNICA' in tipo_doc:
                        if num_doc in Boletas.set_index('Nº Documento').index:
                            debitos.at[idx, 'Marketplace'] = Boletas.loc[Boletas['Nº Documento'] == num_doc, 'Marketplace'].iloc[0]
                    elif 'FACTURA DE EXPORTACIÓN ELECTRÓNICA' in tipo_doc:
                        if num_doc in Exportacion.set_index('Nº Documento').index:
                            debitos.at[idx, 'Marketplace'] = Exportacion.loc[Exportacion['Nº Documento'] == num_doc, 'Marketplace'].iloc[0]
                    elif 'NOTA DE CRÉDITO ELECTRÓNICA' in tipo_doc:
                        if num_doc in notas.set_index('Nº Documento').index:
                            debitos.at[idx, 'Marketplace'] = notas.loc[notas['Nº Documento'] == num_doc, 'Marketplace'].iloc[0]

            # Mapear Marketplace para notas de débito
            df1['Numero Documento'] = df1['Numero Documento'].astype(str).str.strip()
            debitos['Nº Documento'] = debitos['Nº Documento'].astype(str).str.strip()
            diccionario_debitos = dict(zip(debitos['Nº Documento'], debitos['Marketplace']))
            condicion_debito = df1['Tipo de Documento'].str.contains('NOTA DE DÉBITO ELECTRÓNICA', case=False)
            df1.loc[condicion_debito, 'Marketplace'] = df1.loc[condicion_debito, 'Numero Documento'].map(diccionario_debitos)

            # Después de procesar todos los tipos de documentos y antes de crear las tablas dinámicas
            # Rellenar los valores vacíos o NaN en la columna Marketplace con "Sin datos"
            df1['Marketplace'] = df1['Marketplace'].fillna('Sin datos')
            df1['Marketplace'] = df1['Marketplace'].replace(['', None], 'Sin datos')



            # Modificar la parte de filtrado
            df_filtrado = df1.copy()
            if filtros_columnas and filtros_valores:
                try:
                    for columna, valores in zip(filtros_columnas, filtros_valores):
                        # Crear máscara para excluir los valores específicos
                        mascara = ~df_filtrado[columna].isin(valores)
                        df_filtrado = df_filtrado[mascara]
                except Exception as e:
                    messagebox.showwarning("Advertencia", f"Error al aplicar filtros: {e}\nSe continuará sin filtros.")
                    df_filtrado = df1.copy()

            # Crear tabla dinámica de Sucursal
            Sucursal = df_filtrado.pivot_table(index='Sucursal', aggfunc={'Cantidad': 'sum', 'Subtotal Bruto': 'sum','Subtotal Impuestos':'sum'}).round(0).astype(int)

            # Tabla dinamica para ventas y costos
            df_filtrado['Total venta'] = df_filtrado['Subtotal Bruto'] - df_filtrado['Subtotal Impuestos']
            Ventas = df_filtrado.pivot_table(index='Tipo de Producto / Servicio', aggfunc={'Total venta': 'sum', 'Costo Total Neto': 'sum'})
            Total_ventas = df_filtrado['Total venta'].sum()
            Total_costos = df_filtrado['Costo Total Neto'].sum()
            Ventas.loc['Total'] = [Total_costos, Total_ventas]
            
            # Calcular margen para Ventas
            Ventas['Margen (%)'] = Ventas.apply(
                lambda row: calcular_margen(row['Total venta'], row['Costo Total Neto']), 
                axis=1
            ).round(2).astype('int64')
            
            # Calcular los que tienen margen menor o igual a 20% para Ventas
            Ventas_menor_margen = Ventas.loc[Ventas['Margen (%)'] <= 20, ['Total venta', 'Costo Total Neto', 'Margen (%)']]

            # Ahora que tenemos todos los datos de Marketplace, creamos las tablas dinámicas
            Marketplace = df_filtrado.pivot_table(index='Marketplace', aggfunc={'Total venta': 'sum', 'Costo Total Neto': 'sum'})

            Total_ventas_marketplace = Marketplace['Total venta'].sum()
            Total_costos_marketplace = Marketplace['Costo Total Neto'].sum()
            Marketplace.loc['Total'] = [Total_costos_marketplace, Total_ventas_marketplace]

            # Calcular margen para Marketplace
            Marketplace['Margen (%)'] = Marketplace.apply(
                lambda row: calcular_margen(row['Total venta'], row['Costo Total Neto']), 
                axis=1
            ).round(2).astype('int64')
            
            # Calcular los que tienen margen menor o igual a 20% para Marketplace
            Marketplace_menor_margen = Marketplace.loc[Marketplace['Margen (%)'] <= 20, ['Total venta', 'Costo Total Neto', 'Margen (%)']]

            # Calcular los productos más vendidos y menos vendidos, ahora agrupando por 'Tipo de Producto / Servicio' y 'Producto'.
            agrupados = df_filtrado.groupby(['Tipo de Producto / Servicio', 'Producto'])['Cantidad'].sum().sort_values(ascending=False)
            top_10 = agrupados.head(10).reset_index(name='Cantidad mas vendidos')
            bottom_10 = agrupados.tail(10).reset_index(name='Cantidad menos vendidos')



        # Guardar el archivo final
        output_file = filedialog.asksaveasfilename(
            title="Guardar archivo",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not output_file:
            return

        # Modificar la sección de guardado de archivos
        output_file2 = os.path.join(os.path.dirname(output_file), "output_file2.xlsx")

        with pd.ExcelWriter(output_file2) as writer2:
            Boletas.to_excel(writer2, sheet_name='Boletas', index=False, startrow=2)
            notas.to_excel(writer2, sheet_name='Notas de Crédito', index=False, startrow=2)
            debitos.to_excel(writer2, sheet_name='Notas de Débito', index=False, startrow=2)
            facturas.to_excel(writer2, sheet_name='Facturas', index=False, startrow=2)
            Exportacion.to_excel(writer2, sheet_name='Facturas Exportación', index=False, startrow=2)
            DESPACHO.to_excel(writer2, sheet_name='Guías de Despacho', index=False, startrow=2)

        # Cargar el archivo para aplicar formato
        workbook2 = load_workbook(output_file2)

        def aplicar_estilos_hoja(hoja):
            # Aplica formato a la cabecera
            for cell in hoja[3]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000")
                )
            # Ajustar altura y ancho de columnas
            for row in hoja.iter_rows(min_row=3):
                hoja.row_dimensions[row[0].row].height = 20
            for col in hoja.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                hoja.column_dimensions[col_letter].width = max_length + 2
            # Destacar la fila "Total"
            for row in hoja.iter_rows(min_row=4):
                if row[0].value == 'Total':
                    for cell in row:
                        cell.font = Font(bold=True, color="000000")
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Reemplazar cada bloque repetido:
        for sheet_name in ['Boletas', 'Notas de Crédito', 'Notas de Débito', 
                        'Facturas', 'Facturas Exportación', 'Guías de Despacho']:
            aplicar_estilos_hoja(workbook2[sheet_name])

        # Guardar el archivo con los cambios
        workbook2.save(output_file2)
        workbook2.close()

        output_file = filedialog.asksaveasfilename(
            title="Guardar archivo final",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not output_file:
            exit()

        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df1.to_excel(writer, sheet_name='Actualizado', startrow=2)
            Sucursal.to_excel(writer, sheet_name='total ventas', startrow=2)
            Ventas.to_excel(writer, sheet_name='vendido', startrow=2)
            start_row = len(Ventas) + 4
            Marketplace.to_excel(writer, sheet_name='vendido', startrow=start_row)
            Ventas_menor_margen.to_excel(writer, sheet_name='vendido', startrow=start_row + len(Marketplace) + 4)
            Marketplace_menor_margen.to_excel(writer, sheet_name='vendido', startrow=start_row + len(Marketplace) + 4 + len(Ventas_menor_margen) + 4)
            top_10.to_excel(writer, sheet_name='total ventas', startrow= 2 + len(Sucursal) + 4)
            bottom_10.to_excel(writer, sheet_name='total ventas', startrow= 2 + len(Sucursal) + 4 + len(top_10) + 4)


        # Cargar el archivo para aplicar formato
        workbook = load_workbook(output_file)






        # Reemplazar cada bloque repetido:
        for sheet_name in ['Actualizado', 'total ventas', 'vendido']:
            if sheet_name in workbook.sheetnames:
                aplicar_estilos_hoja(workbook[sheet_name])

        # Guardar el archivo con los cambios
        workbook.save(output_file)
        workbook.close()

    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error al ejecutar el programa: {e}")

# Crear la interfaz gráfica
root = Tk()
root.title("Generador Excel")
root.configure(bg='#f0f0f0')  # Color de fondo suave

# Configurar el estilo
style = ttk.Style()
style.theme_use('clam')
style.configure('Custom.TButton',
    background='#4a90e2',
    foreground='white',
    padding=10,
    font=('Helvetica', 10, 'bold')
)
style.configure('Title.TLabel',
    font=('Helvetica', 14, 'bold'),
    background='#f0f0f0',
    foreground='#2c3e50'
)
style.configure('Subtitle.TLabel',
    font=('Helvetica', 10),
    background='#f0f0f0',
    foreground='#34495e'
)

# Crear frame principal
main_frame = ttk.Frame(root, padding="20")
main_frame.pack(fill='both', expand=True)

# Título principal
label_instrucciones = ttk.Label(
    main_frame,
    text="Generador de Reportes Excel",
    style='Title.TLabel'
)
label_instrucciones.pack(pady=20)

# Frame para selección de archivos
files_frame = ttk.LabelFrame(main_frame, text="Selección de Archivos", padding="10")
files_frame.pack(fill='x', padx=20, pady=10)

# Botones de selección de archivos con sus etiquetas
for button_text, command, label_var in [
    ("Archivo de Ventas", seleccionar_archivo_ventas, "label_ventas"),
    ("Archivo de Costos", seleccionar_archivo_costos, "label_costos"),
    ("Archivo del Año", seleccionar_archivo_ano, "label_ano")
]:
    file_frame = ttk.Frame(files_frame)
    file_frame.pack(fill='x', pady=5)
    
    btn = ttk.Button(
        file_frame,
        text=button_text,
        command=command,
        style='Custom.TButton'
    )
    btn.pack(side='left', padx=5)
    
    globals()[label_var] = ttk.Label(
        file_frame,
        text="No seleccionado",
        style='Subtitle.TLabel'
    )
    globals()[label_var].pack(side='left', padx=5)

# Frame para configuración
config_frame = ttk.LabelFrame(main_frame, text="Configuración", padding="10")
config_frame.pack(fill='x', padx=20, pady=10)

# Entradas de configuración
entries = [
    ("Fila de inicio de ventas mensuales:", "fila_ventas_entry"),
    ("Fila de inicio de ventas anuales:", "fila_ano_entry")
]

for label_text, entry_var in entries:
    entry_frame = ttk.Frame(config_frame)
    entry_frame.pack(fill='x', pady=5)
    
    label = ttk.Label(
        entry_frame,
        text=label_text,
        style='Subtitle.TLabel'
    )
    label.pack(side='left', padx=5)
    
    globals()[entry_var] = ttk.Entry(entry_frame)
    globals()[entry_var].pack(side='right', padx=5, fill='x', expand=True)

# Agregar después del config_frame:
filtros_tablas_frame = ttk.LabelFrame(main_frame, text="Filtros para Tablas Dinámicas", padding="10")
filtros_tablas_frame.pack(fill='x', padx=20, pady=10)

# Crear entradas para filtros
filtros_columnas_frame = ttk.Frame(filtros_tablas_frame)
filtros_columnas_frame.pack(fill='x', pady=5)
ttk.Label(
    filtros_columnas_frame, 
    text="Columnas a filtrar (separadas por comas):",
    style='Subtitle.TLabel'
).pack(side='left', padx=5)
filtros_columnas_entry = ttk.Entry(filtros_columnas_frame)
filtros_columnas_entry.pack(side='right', padx=5, fill='x', expand=True)

filtros_valores_frame = ttk.Frame(filtros_tablas_frame)
filtros_valores_frame.pack(fill='x', pady=5)
ttk.Label(
    filtros_valores_frame, 
    text="Valores a excluir (usar ; entre valores y , entre columnas):",
    style='Subtitle.TLabel'
).pack(side='left', padx=5)
filtros_valores_entry = ttk.Entry(filtros_valores_frame)
filtros_valores_entry.pack(side='right', padx=5, fill='x', expand=True)

# Agregar texto de ayuda
ttk.Label(
    filtros_tablas_frame,
    text="Ejemplo: Para filtrar Marketplace y Tipo, escribir:\nColumnas: Marketplace, Tipo de Documento\nValores: vtex;ripley, boleta;factura",
    style='Subtitle.TLabel',
    justify='left'
).pack(pady=5)

button_filtros = ttk.Button(
    filtros_tablas_frame,
    text="Aplicar Filtros",
    command=seleccionar_filtros_tablas,
    style='Custom.TButton'
)
button_filtros.pack(pady=10)

# Frame para botones de acción
action_frame = ttk.Frame(main_frame)
action_frame.pack(fill='x', padx=20, pady=20)

# Botones de acción
button_seleccionar_lineas = ttk.Button(
    action_frame,
    text="Seleccionar Líneas",
    command=seleccionar_lineas,
    style='Custom.TButton'
)
button_seleccionar_lineas.pack(side='left', padx=5, expand=True)

button_ejecutar = ttk.Button(
    action_frame,
    text="Ejecutar",
    command=ejecutar_programa,
    style='Custom.TButton'
)
button_ejecutar.pack(side='right', padx=5, expand=True)

# Centrar la ventana en la pantalla
root.update_idletasks()
width = root.winfo_width()
height = root.winfo_height()
x = (root.winfo_screenwidth() // 2) - (width // 2)
y = (root.winfo_screenheight() // 2) - (height // 2)
root.geometry('{}x{}+{}+{}'.format(width, height, x, y))

root.mainloop()

